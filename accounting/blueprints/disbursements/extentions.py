import os
from openpyxl import Workbook
from pandas import DataFrame
import numpy as np

from .. accounts import Accounts
from .. company import Company


def create_journal(data, app, date_from, date_to):
    list_files = os.listdir(os.path.join(app.instance_path, "temp"))
    for file in list_files:
        os.remove(os.path.join(app.instance_path, "temp", file))

    filename = os.path.join(app.instance_path, "temp", "disbursements journal.xlsx")

    wb = Workbook()

    WriteData(wb, data, date_from, date_to)

    wb.save(filename)
    wb.close()

    return filename


class WriteData:
    def __init__(self, wb, data, date_from, date_to):
        ws = wb["Sheet"]
        ws.title = "Disbursement Journal"

        self.voucher_columns = ["Date", "No.", "Check Number", "Vendor", "Particulars"]

        reformed_data = self.reform_data(data)
        row_start = row_num = self.write_headers(ws, date_from, date_to)
        row_num = self.write_details(ws, reformed_data, row_num)

        column_start = len(self.voucher_columns) + 1
        self.write_footer(ws, row_num, row_start, column_start)

    def reform_data(self, data):
        df_data = DataFrame(columns=self.voucher_columns)

        for voucher in data:
            _dict = {
                "Date": voucher.record_date,
                "No.": voucher.disbursement_number,
                "Check Number": voucher.check_number,
                "Vendor": voucher.vendor.vendor_name,
                "Particulars": voucher.notes
            }

            for entry in voucher.entries:
                account_title = entry.account.account_title
                if account_title not in df_data.columns:
                    df_data[account_title] = 0
                if account_title not in _dict:
                    _dict[account_title] = 0
                _dict[account_title] += entry.debit - entry.credit

            df_data = df_data.append(_dict, ignore_index=True)

        preferred_accounts = [account.account_title for account in Accounts.query.filter(
                                        Accounts.account_type_id == 1
                                        ).order_by(
                                                Accounts.account_number
                                                ).all()
                                        if account.account_title in df_data.columns
                               ]

        other_accounts = Accounts.query.filter(~Accounts.account_title.in_(preferred_accounts)
                                               ).order_by(Accounts.account_number).all()

        other_accounts = [account.account_title for account in other_accounts
                          if account.account_title in df_data.columns]

        self.accounts = preferred_accounts + other_accounts
        df_data = df_data[self.voucher_columns + self.accounts]
        df_data = df_data.replace(0, np.nan)

        return df_data

    @staticmethod
    def write_headers(ws, date_from, date_to):
        if date_from.year != date_to.year:
            date_from = date_from.strftime('%B %-d, %Y')
            date_to = date_to.strftime('%B %-d, %Y')
            date_range = f"From {date_from} to {date_to}"

        elif date_from.month != date_to.month:
            date_from = date_from.strftime('%B %-d')
            date_to = date_to.strftime('%B %-d, %Y')
            date_range = f"From {date_from} to {date_to}"

        elif date_from == date_to:
            date_from = date_from.strftime('%B %-d, %Y')
            date_range = f"For {date_from}"

        else:
            date_from = date_from.strftime('%B %-d')
            date_to = date_to.strftime('%-d, %Y')
            date_range = f"From {date_from} to {date_to}"

        row_num = 1
        ws[f"A{row_num}"].value = Company.query.get(1).company_name

        row_num += 1
        ws[f"A{row_num}"].value = "Disbursement Journal"

        row_num += 1
        ws[f"A{row_num}"].value = date_range

        row_num += 2

        return row_num

    @staticmethod
    def write_details(ws, reformed_data, row_num):
        column_ref = {value: i + 1 for i, value in enumerate(reformed_data.columns)}
        for value, i in column_ref.items():
            cell = ws.cell(row=row_num, column=i)
            cell.value = value
        row_num += 1

        for i, row in reformed_data.iterrows():
            for key, column in column_ref.items():
                cell = ws.cell(row=row_num, column=column)
                cell.value = row[key]
            row_num += 1

        row_num += 1

        return row_num

    def write_footer(self, ws, row_num, row_start, column_start):
        row_end = row_num - 1

        cell = ws[f"A{row_num}"]
        cell.value = "Total"

        for column in range(column_start, column_start + len(self.accounts)):
            cell = ws.cell(row=row_num, column=column)
            cell.value = f"=SUM({cell.column_letter}{row_start}:{cell.column_letter}{row_end})"
