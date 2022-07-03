import os
from pandas import DataFrame, concat
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side

from .. account_type import AccountType
from .. accounts import Accounts
from .. company import Company

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

double_rule_border = Border(bottom=Side(style='double'))

ALIGNMENT = {
                "Date": Alignment(horizontal="center", vertical="top"),
                "No.": Alignment(horizontal="center", vertical="top"),
                "Invoice Number": Alignment(horizontal="center", vertical="top"),
                "Vendor": Alignment(horizontal="left", vertical="top", wrap_text=True),
                "Particulars": Alignment(horizontal="left", vertical="top", wrap_text=True)
            }

NUMBER_FORMAT = {
                "Date": "yyyy-mmm-dd",
                "No.": "General",
                "Invoice Number": "General",
                "Vendor": "General",
                "Particulars": "General"
            }

COLUMN_WIDTH = {
                "Date": 12,
                "No.": 10,
                "Invoice Number": 12,
                "Vendor": 20,
                "Particulars": 25
            }


def create_journal(data, app, date_from, date_to):
    list_files = os.listdir(os.path.join(app.instance_path, "temp"))
    for file in list_files:
        os.remove(os.path.join(app.instance_path, "temp", file))

    filename = os.path.join(app.instance_path, "temp", "accounts payable journal.xlsx")

    wb = Workbook()

    WriteData(wb, data, date_from, date_to)

    wb.save(filename)
    wb.close()

    return filename


class WriteData:
    def __init__(self, wb, data, date_from, date_to):
        ws = wb["Sheet"]
        ws.title = "APJ"

        self.voucher_columns = ["Date", "No.", "Invoice Number", "Vendor", "Particulars"]

        reformed_data = self.reform_data(data)
        row_start = row_num = self.write_headers(ws, date_from, date_to)
        row_num = self.write_details(ws, reformed_data, row_num)

        column_start = len(self.voucher_columns) + 1
        self.write_footer(ws, row_num, row_start, column_start)

    def reform_data(self, data):
        df_data = DataFrame(columns=self.voucher_columns)

        for voucher in data:
            _dict = {
                "Date": voucher.record_date,
                "No.": voucher.accounts_payable_number,
                "Invoice Number": voucher.invoice_number,
                "Vendor": voucher.vendor.vendor_name,
                "Particulars": voucher.notes
            }

            for entry in voucher.entries:
                account_title = entry.account.account_title
                if account_title not in df_data.columns:
                    df_data[account_title] = 0
                if account_title not in _dict:
                    _dict[account_title] = 0
                _dict[account_title] += entry.debit - entry.credit

            df_data = concat([df_data, DataFrame(_dict, index={len(df_data)+1})])

        account_type = AccountType.query.filter_by(account_type="Accounts Payable").first()
        preferred_accounts = [account.account_title for account in Accounts.query.filter(
                                        Accounts.account_type_id == account_type.id
                                        ).order_by(
                                                Accounts.account_number
                                                ).all()
                                        if account.account_title in df_data.columns
                               ]

        other_accounts = Accounts.query.filter(~Accounts.account_title.in_(preferred_accounts)
                                               ).order_by(Accounts.account_number).all()

        other_accounts = [account.account_title for account in other_accounts
                          if account.account_title in df_data.columns]

        self.accounts = preferred_accounts + other_accounts
        df_data = df_data[self.voucher_columns + self.accounts]
        df_data = df_data.replace(0, np.nan)

        return df_data

    @staticmethod
    def write_headers(ws, date_from, date_to):
        if date_from.year != date_to.year:
            date_from = date_from.strftime('%B %d, %Y')
            date_to = date_to.strftime('%B %d, %Y')
            date_range = f"From {date_from} to {date_to}"

        elif date_from.month != date_to.month:
            date_from = date_from.strftime('%B %d')
            date_to = date_to.strftime('%B %d, %Y')
            date_range = f"From {date_from} to {date_to}"

        elif date_from == date_to:
            date_from = date_from.strftime('%B %d, %Y')
            date_range = f"For {date_from}"

        else:
            date_from = date_from.strftime('%B %d')
            date_to = date_to.strftime('%d, %Y')
            date_range = f"From {date_from} to {date_to}"

        row_num = 1
        cell = ws[f"A{row_num}"]
        cell.value = Company.query.get(1).company_name
        cell.font = Font(size=14, bold=True)

        row_num += 1
        cell = ws[f"A{row_num}"]
        cell.value = "Accounts Payable Journal"
        cell.font = Font(size=10, bold=True)

        row_num += 1
        cell = ws[f"A{row_num}"]
        cell.value = date_range
        cell.font = Font(size=10, bold=True)

        row_num += 2

        return row_num

    @staticmethod
    def write_details(ws, reformed_data, row_num):
        column_ref = {value: i + 1 for i, value in enumerate(reformed_data.columns)}
        for value, i in column_ref.items():
            cell = ws.cell(row=row_num, column=i)

            # Adjust column width
            column_letter = cell.column_letter
            ws.column_dimensions[column_letter].width = 14 if value not in COLUMN_WIDTH else COLUMN_WIDTH[value]

            # Write column header
            cell.value = value
            cell.font = Font(size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
        row_num += 1

        # Write row details
        for i, row in reformed_data.iterrows():
            for key, column in column_ref.items():
                cell = ws.cell(row=row_num, column=column)
                cell.value = row[key]
                cell.font = Font(size=10)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right", vertical="top") \
                    if key not in ALIGNMENT else ALIGNMENT[key]
                cell.number_format = "#,##0.00; (#,##0.00)" if key not in NUMBER_FORMAT else NUMBER_FORMAT[key]
            row_num += 1

        row_num += 1

        return row_num

    def write_footer(self, ws, row_num, row_start, column_start):
        row_end = row_num - 1

        cell = ws[f"A{row_num}"]
        cell.value = "Total"
        cell.font = Font(size=10, bold=True)

        for column in range(column_start, column_start + len(self.accounts)):
            cell = ws.cell(row=row_num, column=column)
            cell.value = f"=SUM({cell.column_letter}{row_start}:{cell.column_letter}{row_end})"
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="right")
            cell.border = double_rule_border
            cell.number_format = "#,##0.00; (#,##0.00)"
